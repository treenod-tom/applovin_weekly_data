import requests
from urllib.request import urlretrieve
import json
import pandas as pd
import numpy as np
import os
import datetime
import pytz
import sys

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import warnings
warnings.filterwarnings('ignore')

def api_access(date, platform):
    aggregated = 'false'
    # 환경변수에서 API 키 가져오기
    api_key = os.environ.get('APPLOVIN_API_KEY')
    if not api_key:
        print("❌ API 키가 설정되지 않았습니다!")
        sys.exit(1)
        
    application = 'com.linecorp.lgpkpk'
    url = "https://r.applovin.com/max/userAdRevenueReport?api_key={}&date={}&platform={}&application={}&aggregated={}"
    
    url = url.format(api_key, date, platform, application, aggregated)
    
    response = requests.get(url)
    
    if response.status_code == 200:
        return response.json()['ad_revenue_report_url'] 
    else:
        print(f"❌ 오류 발생: 상태 코드 {response.status_code}") 
        sys.exit(1)

def get_download(url, file_name, platform, directory):
    try:
        # 현재 작업 디렉토리 확인 (GitHub Actions는 /github/workspace)
        print(f"현재 작업 디렉토리: {os.getcwd()}")
        
        # 디렉토리가 없으면 생성
        os.makedirs(directory, exist_ok=True)
        
        # 파일 경로 생성 (Linux 경로 사용)
        file_path = os.path.join(directory, f'{file_name}_{platform}_raw.csv')
        
        urlretrieve(url, file_path)
        print(f'{file_name} {platform} raw 파일 다운로드 완료: {file_path}')
        return file_path
        
    except Exception as e:
        print(f'다운로드 오류: {e}')
        sys.exit(1)

def data_processing(file_name, date, platform, directory):
    try:
        df = pd.read_csv(file_name)
        bins = np.append(np.arange(0, 31, 1), np.inf)
        labels = [f"{round(i,3)} ~ {round(i+1-0.01,3)}" for i in np.arange(0.0, 30, 1)] + ["30+ "]
        
        df['CPM'] = pd.cut(round(df['Revenue']*1000,3), bins=bins, labels=labels, right=False)

        # CPM별 집계
        cpm_count = df.groupby('CPM').size().reset_index(name='Impressions')
        total_value = cpm_count["Impressions"].sum()
        cpm_count.loc[31] = ['Total', total_value]
        
        # 네트워크별 집계
        network_count = df.groupby(["CPM", "Network"]).size().reset_index(name="Count")
        
        # 결과 파일 경로 (Linux 경로 사용)
        result_file_name = os.path.join(directory, f'{date}_{platform}.xlsx')

        with pd.ExcelWriter(result_file_name, engine="openpyxl") as writer:
            cpm_count.to_excel(writer, sheet_name="CPM", index=False)     
            network_count.to_excel(writer, sheet_name="Price Points", index=False)

            workbook = writer.book

            # CPM 시트 스타일링
            sheet = workbook["CPM"]
            blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            
            # 헤더 스타일
            for cell in sheet[1]:
                cell.fill = blue_fill
            
            # 마지막 행 스타일
            last_row = sheet.max_row
            for cell in sheet[last_row]:
                cell.fill = blue_fill
                cell.font = Font(bold=True)
            
            # 데이터 막대
            if len(cpm_count["Impressions"]) > 1:
                min_value = min(cpm_count["Impressions"][:-1])
                max_value = max(cpm_count["Impressions"][:-1])
                
                data_bar = DataBarRule(start_type="num", start_value=min_value, 
                                     end_type="num", end_value=max_value, 
                                     color="538DD5", showValue="None")
                
                impression_col = f"B2:B{last_row-1}"
                sheet.conditional_formatting.add(impression_col, data_bar)

            # 열 너비 조정
            for col in range(1, sheet.max_column + 1):
                col_letter = get_column_letter(col) 
                sheet.column_dimensions[col_letter].width = 12  

            # 숫자 포맷
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row=row, column=2).number_format = "#,##0"

            # Price Points 시트 스타일링
            sheet = workbook["Price Points"]
            header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=1, column=col).fill = header_fill   
               
            # 셀 병합 로직
            previous_value = None
            start_row = 2
            for row in range(2, sheet.max_row + 1):
                current_value = sheet.cell(row=row, column=1).value
            
                if current_value == previous_value:
                    sheet.cell(row=row, column=1, value="")
                else:
                    if previous_value is not None and start_row < row - 1:
                        sheet.merge_cells(start_row=start_row, end_row=row - 1, start_column=1, end_column=1)
                        sheet.cell(row=start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
            
                    previous_value = current_value
                    start_row = row
            
            # 마지막 병합
            if previous_value is not None and start_row < sheet.max_row:
                sheet.merge_cells(start_row=start_row, end_row=sheet.max_row, start_column=1, end_column=1)
                sheet.cell(row=start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
                
            # 열 너비
            column_widths = {1: 13, 2: 34, 3: 9}
            for col_idx, width in column_widths.items():
                col_letter = get_column_letter(col_idx)
                sheet.column_dimensions[col_letter].width = width

            # 숫자 포맷
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row=row, column=3).number_format = "#,##0"
                
            workbook.save(result_file_name)
            workbook.close()

        print(f'작업파일 저장 완료: {result_file_name}\n')
        return result_file_name
        
    except Exception as e:
        print(f'데이터 처리 오류: {e}')
        sys.exit(1)

def send_mail(file_list, date_list):
    # 환경변수에서 이메일 정보 가져오기
    user_email = os.environ.get('USER_EMAIL')
    app_password = os.environ.get('APP_PASSWORD')
    to_email = os.environ.get('TO_EMAIL')
    cc_email_str = os.environ.get('CC_EMAIL', '')
    
    if not all([user_email, app_password, to_email]):
        print("❌ 이메일 설정이 누락되었습니다!")
        return
    
    cc_email = [email.strip() for email in cc_email_str.split(',') if email.strip()]

    # 이메일 내용
    subject = "[데이터전략팀] AppLovin API 추출 데이터 전달"
    body = f"안녕하세요 하딘 :) \n\n {', '.join(date_list)} AppLovin API 추출 데이터 전달 드립니다! \n 궁금하신점 있으시면 언제든 말씀해주세요 \n\n 감사합니다. \n 톰 드림"
      
    # 이메일 메시지 생성
    msg = MIMEMultipart()
    msg["From"] = user_email
    msg["To"] = to_email
    if cc_email:
        msg["Cc"] = ", ".join(cc_email)
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    # 파일 첨부
    for file_path in file_list:
        if os.path.exists(file_path):
            with open(file_path, "rb") as attachment:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", f"attachment; filename={os.path.basename(file_path)}")
                msg.attach(part)
            print(f"✅ 파일 첨부: {os.path.basename(file_path)}")
        else:
            print(f"⚠️ 파일을 찾을 수 없습니다: {file_path}")

    # 이메일 전송
    all_recipients = [to_email] + cc_email 
    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(user_email, app_password)
        server.sendmail(user_email, all_recipients, msg.as_string())
        server.quit()
        print("✅ 이메일 전송 완료!")
    except Exception as e:
        print(f"❌ 이메일 전송 실패: {e}")

def main():
    # 🕐 시작 시간 기록
    kst = pytz.timezone('Asia/Seoul')
    start_time = datetime.datetime.now(kst)
    print("🚀 AppLovin API 데이터 추출 시작!")
    print(f"⏰ 시작 시간: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    
    # 작업 디렉토리 설정 (현재 디렉토리 기준)
    current_date = datetime.datetime.now(kst).strftime('%Y-%m-%d')
    raw_save_directory = f'raw/{current_date}'
    result_save_directory = f'result/{current_date}'
    
    # 디렉토리 생성
    os.makedirs(raw_save_directory, exist_ok=True)
    os.makedirs(result_save_directory, exist_ok=True)
    
    date_list = []
    file_list = []
    
    # 날짜 설정 (-4일, -2일)
    for i in [-4, -2]:
        target_date = datetime.datetime.now(kst) + datetime.timedelta(days=i)
        date_list.append(target_date.strftime('%Y-%m-%d'))
    
    print(f"📅 처리할 날짜: {date_list}")
    
    # 각 날짜와 플랫폼별로 처리
    for date in date_list:
        for platform in ['ios', 'android']:
            try:
                step_start = datetime.datetime.now(kst)
                print(f"\n📱 처리 중: {date} - {platform}")
                
                # API 호출
                url = api_access(date, platform)
                
                # 파일 다운로드
                raw_file = get_download(url, date, platform, raw_save_directory)
                
                # 데이터 처리
                result_file = data_processing(raw_file, date, platform, result_save_directory)
                file_list.append(result_file)
                
                step_end = datetime.datetime.now(kst)
                step_duration = (step_end - step_start).total_seconds()
                print(f"✅ {date} {platform} 완료 (소요시간: {step_duration:.1f}초)")
                
            except Exception as e:
                print(f"❌ {date} {platform} 처리 실패: {e}")
                continue
    
    # 이메일 전송
    email_start = datetime.datetime.now(kst)
    if file_list:
        print(f"\n📧 이메일 전송 시작...")
        send_mail(file_list, date_list)
        email_end = datetime.datetime.now(kst)
        email_duration = (email_end - email_start).total_seconds()
        print(f"📧 이메일 전송 완료 (소요시간: {email_duration:.1f}초)")
        print(f"\n✅ 총 {len(file_list)}개 파일 처리 완료!")
    else:
        print("❌ 처리된 파일이 없습니다.")
    
    # 🏁 완료 시간 및 총 소요시간 출력
    end_time = datetime.datetime.now(kst)
    total_duration = (end_time - start_time).total_seconds()
    
    print("=" * 60)
    print("🏁 작업 완료!")
    print(f"⏰ 완료 시간: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"⏱️  총 소요시간: {total_duration:.1f}초 ({total_duration/60:.1f}분)")
    print(f"📊 처리 통계:")
    print(f"   - 성공한 파일: {len(file_list)}개")
    print(f"   - 처리한 날짜: {len(date_list)}개")
    print(f"   - 처리한 플랫폼: ios, android")
    print("=" * 60)

if __name__ == "__main__":
    main()
